MAX_DIFFERENT_COUNT = 5

from openpyxl import *

def compareExcel(ename1, ename2):

    print "------------------------------------"
    print "Comparing",ename1,ename2

    wb1 = load_workbook(filename = ename1)
    wb2 = load_workbook(filename = ename2)
    sn1 = wb1.get_sheet_names()
    sn2 = wb2.get_sheet_names()
    print ename1,"has sheet names:", sn1
    print ename2,"has sheet names:", sn2

    if (sn1 != sn2):
        print "Two file has different sheets."
    else:
        print "Two file has the same sheet name."
    print

    sn = sn1
    for wsn in sn:
        ws1 = wb1.get_sheet_by_name(name = wsn)
        ws2 = wb2.get_sheet_by_name(name = wsn)
        c = ws1.get_highest_column()
        r = ws1.get_highest_row()
        if ((ws2.get_highest_column()!= c) or (ws2.get_highest_row() != r)):
            print "DIFFERDENT at SHEET-",wsn,": Rows or columns not the same!"
        else:
            flag = True
            for i in range(1,r+1):
                for j in range(1,c+1):
                    c1 = ws1.cell(None,i,j)
                    c2 = ws2.cell(None,i,j)
                    if (c1):
                        if (c2):
                            if (c1.value != c2.value):
                                if ((wsn == "Internal Info") and (i == 4) and (j == 2)):
                                    continue
                                print "DIFFERDENT_VALUE at SHEET-",wsn,": At (",i,",",j,")",
                                print "diff FROM",c1.value,"TO",c2.value
                                flag = False
                        else:
                            print "DIFFERDENT_TO_NONE at SHEET-",wsn,": At (",i,",",j,")"
                            print "diff FROM",c1.value
                            flag = False
                    else:
                        if (c2):
                            print "DIFFERDENT_TO_NONE at SHEET-",wsn,": At (",i,",",j,")"
                            print "diff FROM",c2.value
                            flag = False
            if flag:
                print "ARE_THE_SAME at SHEET-",wsn

    print "------------------------------------"
    print

compareExcel("1.xlsx","2.xlsx")
